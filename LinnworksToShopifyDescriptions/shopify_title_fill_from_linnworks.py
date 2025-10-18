import os, sys, time, json, argparse, traceback
from typing import Any, Dict, List, Tuple, Optional, Set
from datetime import datetime
import requests
from dotenv import load_dotenv, find_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv

# =======================
# ENV & CONFIG
# =======================
def load_env():
    env_path = os.getenv("ENV_PATH")
    if env_path and os.path.exists(env_path):
        load_dotenv(env_path)
    else:
        load_dotenv(find_dotenv(usecwd=True))

load_env()

# Linnworks
APP_ID      = (os.getenv("LINNWORKS_APPLICATION_ID") or "").strip()
APP_SECRET  = (os.getenv("LINNWORKS_APPLICATION_SECRET") or "").strip()
GRANT_TOKEN = (os.getenv("LINNWORKS_GRANT_TOKEN") or "").strip()
LINNWORKS_SERVER_OVERRIDE = (os.getenv("LINNWORKS_SERVER_OVERRIDE") or "").strip()
LINNWORKS_AUTH_STYLE      = (os.getenv("LINNWORKS_AUTH_STYLE") or "").strip()  # Bearer|Raw optional

# Shopify
SHOPIFY_STORE_NAME   = (os.getenv("SHOPIFY_STORE_NAME") or "").strip()
SHOPIFY_ACCESS_TOKEN = (os.getenv("SHOPIFY_ACCESS_TOKEN") or "").strip()
API_VERSION          = (os.getenv("API_VERSION") or "2025-01").strip()

# Behaviour
DRY_RUN        = (os.getenv("DRY_RUN","false").lower() in ("1","true","yes"))
FORCE_UPDATE   = (os.getenv("FORCE_UPDATE","false").lower() in ("1","true","yes"))

# Channel
CHANNEL_SOURCE    = (os.getenv("CHANNEL_SOURCE") or "EBAY").strip()
CHANNEL_SUBSOURCE = (os.getenv("CHANNEL_SUBSOURCE") or "EBAY1_UK").strip()

# Pacing
REQUEST_DELAY          = float(os.getenv("REQUEST_DELAY", "0.25"))          # Linnworks
SHOPIFY_REQUEST_DELAY  = float(os.getenv("SHOPIFY_REQUEST_DELAY", "0.5"))   # Shopify

# Files
INPUT_XLSX   = os.getenv("INPUT_XLSX", "sku_list.xlsx")
INPUT_SHEET  = os.getenv("INPUT_SHEET", "Sheet1")
SKU_COLUMN   = os.getenv("SKU_COLUMN", "SKU")
TRACKER_PATH = os.getenv("TRACKER_PATH", "shopify_title_fill_tracker.xlsx")

# CSV change-log
CSV_LOG_PATH = os.getenv("CSV_LOG_PATH", "shopify_title_fill_changes.csv")

# Limits / batching
TITLE_MAX    = int(os.getenv("TITLE_MAX", "255"))      # Title / option value guard
FLUSH_EVERY  = int(os.getenv("FLUSH_EVERY", "50"))     # flush tracker every N rows

# =======================
# HEADERS / TRACKER
# =======================
HEADERS = [
    "Timestamp","SKU","ProductID","OldTitle","NewTitle",
    "Source","SubSource","Status","Note","DryRun"
]

CSV_HEADERS = [
    "Timestamp", "SKU", "ProductID",
    "ShopifyOldTitle", "LinnworksNewTitle",
    "Source", "SubSource", "Status", "Note"
]

def ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def log(msg: str):
    print(msg, flush=True)

def sleep_s(t: float):
    if t > 0:
        time.sleep(t)

def ensure_tracker():
    if not os.path.exists(TRACKER_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "TitleFill"
        ws.append(HEADERS)
        widths = [20,28,14,45,45,10,14,14,60,8]
        for i,w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(TRACKER_PATH)

def append_rows(rows: List[List[Any]]):
    ensure_tracker()
    wb = load_workbook(TRACKER_PATH)
    ws = wb["TitleFill"]
    for r in rows:
        ws.append(r)
    wb.save(TRACKER_PATH)

def ensure_csv_log():
    if not os.path.exists(CSV_LOG_PATH):
        with open(CSV_LOG_PATH, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(CSV_HEADERS)

def append_csv_rows(rows: List[Dict[str, Any]]):
    ensure_csv_log()
    with open(CSV_LOG_PATH, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for r in rows:
            w.writerow([
                r.get("Timestamp", ts()),
                r.get("SKU", ""),
                r.get("ProductID", ""),
                r.get("ShopifyOldTitle", ""),
                r.get("LinnworksNewTitle", ""),
                r.get("Source", CHANNEL_SOURCE),
                r.get("SubSource", CHANNEL_SUBSOURCE),
                r.get("Status", ""),
                r.get("Note", ""),
            ])

def to_row(r: Dict[str,Any]) -> List[Any]:
    return [
        r.get("Timestamp") or ts(),
        r.get("SKU",""),
        r.get("ProductID",""),
        r.get("OldTitle",""),
        r.get("NewTitle",""),
        r.get("Source", CHANNEL_SOURCE),
        r.get("SubSource", CHANNEL_SUBSOURCE),
        r.get("Status",""),
        r.get("Note",""),
        str(DRY_RUN).lower(),
    ]

def safe_title(s: str) -> str:
    s = (s or "").strip()
    return s[:TITLE_MAX]

# =======================
# XLSX INPUT
# =======================
def read_skus_from_excel(path: str, sheet: str, sku_col: str) -> List[str]:
    if not os.path.exists(path):
        raise SystemExit(f"Input Excel not found: {path}")
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet}' not found in {path}")
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    col_index = None
    for i, name in enumerate(header, start=1):
        if (str(name) or "").strip().lower() == sku_col.strip().lower():
            col_index = i
            break
    if not col_index:
        raise SystemExit(f"Column '{sku_col}' not found in first row of {path}/{sheet}")

    skus: List[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[col_index-1]
        if val is None:
            continue
        s = str(val).strip()
        if s:
            skus.append(s)
    wb.close()
    # preserve order but unique
    seen: Set[str] = set()
    out: List[str] = []
    for s in skus:
        if s not in seen:
            out.append(s); seen.add(s)
    return out

# =======================
# LINNWORKS
# =======================
def lw_auth() -> Tuple[str, str]:
    if not (APP_ID and APP_SECRET and GRANT_TOKEN):
        raise SystemExit("Missing Linnworks .env vars LINNWORKS_APPLICATION_*")
    url = "https://api.linnworks.net/api/Auth/AuthorizeByApplication"
    payload = {"ApplicationId": APP_ID, "ApplicationSecret": APP_SECRET, "Token": GRANT_TOKEN}
    r = requests.post(url, json=payload, timeout=40)
    r.raise_for_status()
    data = r.json()
    token  = data.get("Token")
    server = (os.getenv("LINNWORKS_SERVER_OVERRIDE") or
              data.get("Server") or data.get("ServerAddress") or data.get("ServerUrl") or "").rstrip("/")
    if not token or not server:
        raise RuntimeError(f"Linnworks auth response missing Token/Server: {data}")
    return token, server

def lw_make_session_with_probe():
    token, server = lw_auth()
    def probe(style):
        h = {"Authorization": (f"Bearer {token}" if style=="Bearer" else token),
             "Accept":"application/json","Content-Type":"application/json"}
        pr = requests.get(f"{server}/api/Inventory/GetChannels", headers=h, timeout=20)
        return pr.status_code==200, h

    forced = (LINNWORKS_AUTH_STYLE or "").strip()
    if forced in ("Bearer","Raw"):
        ok, h = probe(forced)
        if not ok: raise SystemExit(f"Probe failed with forced style {forced} on {server}")
        s = requests.Session(); s.headers.update(h)
        return s, server, forced

    ok, h = probe("Bearer")
    if ok:
        s = requests.Session(); s.headers.update(h)
        return s, server, "Bearer"
    ok, h = probe("Raw")
    if ok:
        s = requests.Session(); s.headers.update(h)
        return s, server, "Raw"
    raise SystemExit("Auth failed with both Bearer and Raw")

def lw_post(s: requests.Session, server: str, path: str, payload: Dict[str,Any]) -> Any:
    url = f"{server}/api{path}" if not server.endswith("/api") else f"{server}{path}"
    r = s.post(url, data=json.dumps(payload), timeout=50)
    if r.status_code != 200:
        raise RuntimeError(f"LW HTTP {r.status_code} on {path}: {r.text[:300]}")
    return r.json()

def lw_get(s: requests.Session, server: str, path: str, params: Dict[str,Any]) -> Any:
    url = f"{server}/api{path}" if not server.endswith("/api") else f"{server}{path}"
    r = s.get(url, params=params, timeout=50)
    if r.status_code != 200:
        raise RuntimeError(f"LW HTTP {r.status_code} on {path}: {r.text[:300]}")
    return r.json()

def lw_get_stock_ids_by_sku(s: requests.Session, server: str, skus: List[str]) -> Dict[str,str]:
    mapping = {}
    try:
        data = lw_post(s, server, "/Inventory/GetStockItemIdsBySKU", {"request":{"SKUS": skus}})
        for it in (data or {}).get("Items", []):
            sku, sid = it.get("SKU"), it.get("StockItemId")
            if sku and sid:
                mapping[sku]=sid
        return mapping
    except Exception:
        url = f"{server}/api/Inventory/GetStockItemIdsBySKU"
        headers = dict(s.headers); headers["Content-Type"] = "application/x-www-form-urlencoded"
        r = s.post(url, data={"request": json.dumps({"SKUS": skus})}, headers=headers, timeout=50)
        if r.status_code != 200:
            raise RuntimeError(f"LW fallback HTTP {r.status_code}: {r.text[:300]}")
        data = r.json()
        for it in (data or {}).get("Items", []):
            sku, sid = it.get("SKU"), it.get("StockItemId")
            if sku and sid:
                mapping[sku]=sid
        return mapping

def lw_get_item_titles(s: requests.Session, server: str, inventory_item_id: str) -> List[Dict[str,Any]]:
    # Expected to return list of {Source, SubSource, Title}
    return lw_get(s, server, "/Inventory/GetInventoryItemTitles", {"inventoryItemId": inventory_item_id})

def lw_get_item_core(s: requests.Session, server: str, inventory_item_id: str) -> Dict[str,Any]:
    # Fallback to base item title if channel title is missing
    return lw_get(s, server, "/Inventory/GetInventoryItemById", {"id": inventory_item_id})

def lw_pick_channel_title(titles: List[Dict[str,Any]], source: str, subsource: str) -> str:
    for t in titles or []:
        if (t.get("Source","").upper()==source.upper()) and (t.get("SubSource","")==subsource):
            if (t.get("Title") or "").strip():
                return t.get("Title").strip()
    return ""

# =======================
# SHOPIFY
# =======================
def shopify_base() -> str:
    if not SHOPIFY_STORE_NAME or not SHOPIFY_ACCESS_TOKEN:
        raise SystemExit("Missing Shopify .env vars SHOPIFY_STORE_NAME / SHOPIFY_ACCESS_TOKEN")
    return f"https://{SHOPIFY_STORE_NAME}.myshopify.com/admin/api/{API_VERSION}"

def shopify_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Accept": "application/json",
        "Content-Type": "application/json",
    })
    return s

def sh_get(s: requests.Session, path: str, params: Dict[str,Any]=None) -> requests.Response:
    url = f"{shopify_base()}{path}"
    return s.get(url, params=params or {}, timeout=50)

def sh_put(s: requests.Session, path: str, payload: Dict[str,Any]) -> requests.Response:
    url = f"{shopify_base()}{path}"
    return s.put(url, json=payload, timeout=50)

def find_variant_by_sku(s: requests.Session, sku: str,
                        fallback_scan: bool = False,
                        scan_pages: int = 5,  # ~1250 products (5 * 250)
                        ) -> Tuple[Optional[int], Optional[int]]:
    """Return (product_id, variant_id) for exact SKU match.
       If API sku filter misbehaves, optionally fallback-scan products."""
    r = sh_get(s, "/variants.json", {"sku": sku})
    if r.status_code == 429:
        time.sleep(2); return find_variant_by_sku(s, sku, fallback_scan, scan_pages)
    r.raise_for_status()
    arr = r.json().get("variants", [])
    # Verify the API actually returned the right SKU
    for v in arr:
        if (v.get("sku") or "").strip() == sku:
            return int(v["product_id"]), int(v["id"])

    # No exact match via API → optionally fallback scan
    if not fallback_scan:
        return None, None

    # Fallback: scan /products.json pages and check variants client-side
    next_url = f"{shopify_base()}/products.json?limit=250&fields=id,variants"
    pages = 0
    while next_url and pages < scan_pages:
        rr = s.get(next_url, timeout=60)
        if rr.status_code == 429:
            time.sleep(2); continue
        rr.raise_for_status()
        products = rr.json().get("products", [])
        for p in products:
            for v in (p.get("variants") or []):
                if (v.get("sku") or "").strip() == sku:
                    return int(p["id"]), int(v["id"])

        # paginate
        link = rr.headers.get("Link") or rr.headers.get("link")
        next_url = None
        if link and 'rel="next"' in link:
            for part in [x.strip() for x in link.split(",")]:
                if 'rel="next"' in part:
                    left = part.split(";")[0].strip()
                    if left.startswith("<") and left.endswith(">"):
                        next_url = left[1:-1]
        pages += 1

    return None, None


def sh_get_variant(s: requests.Session, variant_id: int) -> Dict[str,Any]:
    r = sh_get(s, f"/variants/{variant_id}.json")
    if r.status_code == 429:
        time.sleep(2); return sh_get_variant(s, variant_id)
    r.raise_for_status()
    return r.json().get("variant", {})

def sh_update_variant_option1(s: requests.Session, variant_id: int, option_value: str):
    payload = {"variant": {"id": variant_id, "option1": option_value}}
    r = sh_put(s, f"/variants/{variant_id}.json", payload)
    if r.status_code == 429:
        time.sleep(2); return sh_update_variant_option1(s, variant_id, option_value)
    if r.status_code >= 400:
        raise RuntimeError(f"Shopify PUT {r.status_code}: {r.text[:300]}")

def sh_get_product(s: requests.Session, product_id: int) -> Dict[str,Any]:
    r = sh_get(s, f"/products/{product_id}.json")
    if r.status_code == 429:
        time.sleep(2); return sh_get_product(s, product_id)
    r.raise_for_status()
    return r.json().get("product", {})

def sh_update_product_title(s: requests.Session, product_id: int, new_title: str):
    payload = {"product": {"id": product_id, "title": new_title}}
    r = sh_put(s, f"/products/{product_id}.json", payload)
    if r.status_code == 429:
        time.sleep(2); return sh_update_product_title(s, product_id, new_title)
    if r.status_code >= 400:
        raise RuntimeError(f"Shopify PUT {r.status_code}: {r.text[:300]}")

# =======================
# CORE PROCESS
# =======================
def chunked(seq: List[str], n: int):
    for i in range(0, len(seq), n):
        yield seq[i:i+n]

def run_process(args):
    skus = read_skus_from_excel(args.excel or INPUT_XLSX, args.sheet or INPUT_SHEET, args.sku_col or SKU_COLUMN)
    if args.only_sku:
        only = set(args.only_sku)
        skus = [s for s in skus if s in only]
    if args.limit:
        skus = skus[:args.limit]

    log(f"Store: {SHOPIFY_STORE_NAME}  API: {API_VERSION}  DRY_RUN={DRY_RUN}  FORCE={args.force or FORCE_UPDATE}")
    log(f"Channel: {CHANNEL_SOURCE}/{CHANNEL_SUBSOURCE}")
    log(f"Input: {args.excel or INPUT_XLSX} [{args.sheet or INPUT_SHEET}] SKU_COL='{args.sku_col or SKU_COLUMN}'")
    log(f"[INPUT] {len(skus)} unique SKU(s) to process.")
    if not skus:
        return

    # Ensure outputs exist & announce
    ensure_tracker()
    ensure_csv_log()
    log(f"[TRACKER] Writing to: {TRACKER_PATH}")
    log(f"[CSV-LOG] Writing to: {CSV_LOG_PATH}")

    lw_sess, lw_server, auth_style = lw_make_session_with_probe()
    log(f"[LW] Auth style={auth_style} server={lw_server}")

    # batch resolve SKUs -> StockItemId
    sku_to_id: Dict[str,str] = {}
    batches = list(chunked(skus, 80))
    for idx, batch in enumerate(batches, start=1):
        log(f"[LW] Resolving StockItemIds batch {idx}/{len(batches)} (size={len(batch)})…")
        try:
            mapping = lw_get_stock_ids_by_sku(lw_sess, lw_server, batch)
            sku_to_id.update(mapping)
            log(f"[LW]  + mapped {len(mapping)}/{len(batch)} (total mapped: {len(sku_to_id)})")
        except Exception as e:
            log(f"[LW] GetStockItemIdsBySKU error for batch {idx}: {e}")
        sleep_s(REQUEST_DELAY)

    shop = shopify_session()
    updated_products: Set[int] = set()   # only used in product mode to prevent multiple writes

    buffer: List[Dict[str,Any]] = []     # for incremental tracker flush
    processed = 0

    for sku in skus:
        processed += 1
        row = {
            "Timestamp": ts(),
            "SKU": sku,
            "ProductID": "",
            "OldTitle": "",
            "NewTitle": "",
            "Source": CHANNEL_SOURCE,
            "SubSource": CHANNEL_SUBSOURCE,
            "Status": "PENDING",
            "Note": "",
        }

        sid = sku_to_id.get(sku)
        if not sid:
            row["Status"]="NOT_FOUND"
            row["Note"]="SKU not found in Linnworks"
            buffer.append(row)
            if len(buffer) >= FLUSH_EVERY:
                append_rows([to_row(r) for r in buffer]); log(f"[TRACKER] Flushed {len(buffer)} rows…"); buffer.clear()
            continue

        # 1) Get channel title from Linnworks (EBAY / EBAY1_UK)
        try:
            titles = lw_get_item_titles(lw_sess, lw_server, sid)
            ch_title = lw_pick_channel_title(titles, CHANNEL_SOURCE, CHANNEL_SUBSOURCE)
        except Exception as e:
            row["Status"]="ERROR"
            row["Note"]=f"GetInventoryItemTitles error: {e}"
            buffer.append(row)
            if len(buffer) >= FLUSH_EVERY:
                append_rows([to_row(r) for r in buffer]); log(f"[TRACKER] Flushed {len(buffer)} rows…"); buffer.clear()
            sleep_s(REQUEST_DELAY)
            continue
        finally:
            sleep_s(REQUEST_DELAY)

        # Fallback to base item title if channel empty
        if not ch_title:
            try:
                core = lw_get_item_core(lw_sess, lw_server, sid)
                base_title = (core.get("Title") or "").strip()
            except Exception as e:
                base_title = ""
                row["Note"] = (row["Note"] + f" | Fallback core title error: {e}").strip(" |")
            ch_title = ch_title or base_title

        if not ch_title:
            row["Status"]="MISSING_TITLE"
            row["Note"] = (row["Note"] + " | No channel/base title in Linnworks").strip(" |")
            buffer.append(row)
            if len(buffer) >= FLUSH_EVERY:
                append_rows([to_row(r) for r in buffer]); log(f"[TRACKER] Flushed {len(buffer)} rows…"); buffer.clear()
            continue

        # 2) Resolve product by SKU in Shopify
        try:
            product_id, variant_id = find_variant_by_sku(
                shop, sku,
                fallback_scan=True,   # turn on fallback scan
                scan_pages=3          # tune as needed
            )

        except Exception as e:
            row["Status"]="ERROR"
            row["Note"]=f"Shopify /variants?sku= error: {e}"
            buffer.append(row)
            if len(buffer) >= FLUSH_EVERY:
                append_rows([to_row(r) for r in buffer]); log(f"[TRACKER] Flushed {len(buffer)} rows…"); buffer.clear()
            sleep_s(SHOPIFY_REQUEST_DELAY)
            continue
        finally:
            sleep_s(SHOPIFY_REQUEST_DELAY)

        if not product_id:
            row["Status"]="SKIPPED"
            row["Note"] = (row["Note"] + " | SKU not found in Shopify").strip(" |")
            buffer.append(row)
            if len(buffer) >= FLUSH_EVERY:
                append_rows([to_row(r) for r in buffer]); log(f"[TRACKER] Flushed {len(buffer)} rows…"); buffer.clear()
            continue

        row["ProductID"] = str(product_id)

        # 3) Compare & update (product title OR variant option1)
        try:
            prod = sh_get_product(shop, product_id)
            old_product_title = (prod.get("title") or "").strip()
            target_text = safe_title(ch_title)

            if not target_text.strip():
                row["Status"]="ERROR"; row["Note"]=(row["Note"]+" | Empty target title").strip(" |")
                row["OldTitle"] = old_product_title
                row["NewTitle"] = target_text

            elif args.variant_option1:
                # ----- VARIANT MODE: per-SKU label via option1 -----
                row["OldTitle"] = old_product_title              # audit reference
                row["NewTitle"] = target_text

                if DRY_RUN or args.dry_run:
                    row["Status"] = "READY"
                    row["Note"] = (row["Note"] + " | DRY_RUN: no write (variant option1)").strip(" |")
                else:
                    # write option1 and verify
                    sh_update_variant_option1(shop, variant_id, target_text)
                    v = sh_get_variant(shop, variant_id)
                    confirmed = (v.get("option1") or "").strip()
                    if confirmed == target_text:
                        row["Status"] = "UPDATED"
                    else:
                        row["Status"] = "VERIFY_FAIL"
                        row["Note"] = (row["Note"] + f" | Variant read-back mismatch: got '{confirmed[:80]}'").strip(" |")

                # CSV log
                append_csv_rows([{
                    "Timestamp": ts(),
                    "SKU": sku,
                    "ProductID": str(product_id),
                    "ShopifyOldTitle": old_product_title,
                    "LinnworksNewTitle": target_text,
                    "Source": CHANNEL_SOURCE,
                    "SubSource": CHANNEL_SUBSOURCE,
                    "Status": row["Status"],
                    "Note": row.get("Note",""),
                }])

            else:
                # ----- PRODUCT MODE: one write per product per run -----
                row["OldTitle"] = old_product_title
                row["NewTitle"] = target_text

                if (not FORCE_UPDATE and not args.force) and old_product_title and (old_product_title == target_text):
                    row["Status"]="SKIPPED"
                    row["Note"] = (row["Note"] + " | Title already matches").strip(" |")
                elif DRY_RUN or args.dry_run:
                    row["Status"] = "READY"
                    row["Note"] = (row["Note"] + " | DRY_RUN: no write").strip(" |")
                else:
                    if product_id in updated_products:
                        row["Status"] = "SKIPPED_DUPLICATE_PRODUCT"
                        row["Note"] = (row["Note"] + " | Product already updated in this run").strip(" |")
                    else:
                        # write title and verify
                        sh_update_product_title(shop, product_id, target_text)
                        confirm = sh_get_product(shop, product_id)
                        confirmed = (confirm.get("title") or "").strip()
                        if confirmed == target_text:
                            row["Status"] = "UPDATED"
                            updated_products.add(product_id)
                        else:
                            row["Status"] = "VERIFY_FAIL"
                            row["Note"] = (row["Note"] + f" | Product read-back mismatch: got '{confirmed[:80]}'").strip(" |")

                # CSV log
                append_csv_rows([{
                    "Timestamp": ts(),
                    "SKU": sku,
                    "ProductID": str(product_id),
                    "ShopifyOldTitle": old_product_title,
                    "LinnworksNewTitle": target_text,
                    "Source": CHANNEL_SOURCE,
                    "SubSource": CHANNEL_SUBSOURCE,
                    "Status": row["Status"],
                    "Note": row.get("Note",""),
                }])

        except Exception as e:
            row["Status"]="ERROR"
            row["Note"]=f"Shopify update error: {e}"
        finally:
            sleep_s(SHOPIFY_REQUEST_DELAY)

        buffer.append(row)
        if len(buffer) >= FLUSH_EVERY:
            append_rows([to_row(r) for r in buffer])
            log(f"[TRACKER] Flushed {len(buffer)} rows… ({processed}/{len(skus)})")
            buffer.clear()

        if processed % 100 == 0:
            log(f"[PROGRESS] {processed}/{len(skus)} processed…")

    # final flush
    if buffer:
        append_rows([to_row(r) for r in buffer])
        log(f"[TRACKER] Final flush {len(buffer)} rows. Total processed: {processed}/{len(skus)}")

    log(f"[DONE] Tracker: {TRACKER_PATH} | CSV log: {CSV_LOG_PATH}")

# =======================
# CLI
# =======================
def parse_args():
    ap = argparse.ArgumentParser(description="Fill Shopify product titles or variant option1 from Linnworks channel titles (eBay/EBAY1_UK) for SKUs listed in an Excel file.")
    ap.add_argument("--excel", help="Path to Excel with SKUs (default from env INPUT_XLSX)")
    ap.add_argument("--sheet", help="Sheet name (default from env INPUT_SHEET)")
    ap.add_argument("--sku-col", help="Column header name for SKU (default from env SKU_COLUMN)")
    ap.add_argument("--only-sku", action="append", help="Limit to specific SKU(s); can repeat")
    ap.add_argument("--limit", type=int, help="Process only first N SKUs")
    ap.add_argument("--force", action="store_true", help="Overwrite even if Shopify already has a different title (product mode)")
    ap.add_argument("--dry-run", action="store_true", help="Do not write to Shopify")
    ap.add_argument("--variant-option1", action="store_true",
                    help="Update variant option1 (per-SKU label) instead of product title")
    return ap.parse_args()

def main():
    args = parse_args()
    global DRY_RUN
    if args.dry_run:
        DRY_RUN = True
    try:
        run_process(args)
    except Exception as e:
        log(f"[FATAL] {e}\n{traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    main()
