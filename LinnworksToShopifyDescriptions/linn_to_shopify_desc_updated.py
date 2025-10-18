"""
===============================================================================
🧩  SCRIPT NAME:  linn_to_shopify_desc.py
===============================================================================
PURPOSE:
    Fetch product descriptions from Linnworks and upload them to your Shopify
    store for products that currently have an EMPTY description.

    - Reads SKUs from Shopify (via Admin API)
    - Pulls matching descriptions from Linnworks (by SKU → StockItemId)
    - Updates Shopify product `body_html` only if it’s empty
    - Tracks all steps in Excel: shopify_desc_fill_tracker.xlsx
    - Supports dry-run, partial reruns, verification, and resumable operation.

===============================================================================
⚙️  ENVIRONMENT VARIABLES (.env EXAMPLE)
===============================================================================
LINNWORKS_APPLICATION_ID=xxxxxxxxxxxxxxxx
LINNWORKS_APPLICATION_SECRET=xxxxxxxxxxxxxxxx
LINNWORKS_GRANT_TOKEN=xxxxxxxxxxxxxxxx

SHOPIFY_STORE_NAME=yourshopname                # Only subdomain (no ".myshopify.com")
SHOPIFY_ACCESS_TOKEN=shpat_xxxxxxxxxxxxxxxx
API_VERSION=2025-01

DRY_RUN=false                                  # true → test mode (no updates)
FORCE_UPDATE=false                             # true → overwrite non-empty desc

CHANNEL_SOURCE=EBAY                            # Linnworks channel source
CHANNEL_SUBSOURCE=EBAY1_UK                     # Linnworks channel subsource

REQUEST_DELAY=0.2                              # Delay between Linnworks requests
SHOPIFY_REQUEST_DELAY=0.5                      # Delay between Shopify requests
TRACKER_PATH=shopify_desc_fill_tracker.xlsx
STATE_PATH=.linnworks_shopify_desc.state.json

===============================================================================
🧭  AVAILABLE MODES
===============================================================================
--mode discover     → Find Shopify products with empty description.
--mode populate     → Fetch Linnworks descriptions for discovered SKUs.
--mode update       → Upload Linnworks HTML to Shopify.
--mode verify       → Check that Shopify descriptions were successfully updated.
--mode process      → Run all steps sequentially (discover → populate → update → verify).

===============================================================================
🏃  HOW TO RUN
===============================================================================
1️⃣ Discover empty Shopify descriptions:
    python linn_to_shopify_desc.py --mode discover

2️⃣ Populate Linnworks descriptions:
    python linn_to_shopify_desc.py --mode populate

3️⃣ Update Shopify with new descriptions:
    python linn_to_shopify_desc.py --mode update

4️⃣ Verify updates (re-check Shopify):
    python linn_to_shopify_desc.py --mode verify

5️⃣ Run all in one go (auto pipeline):
    python linn_to_shopify_desc.py --mode process

===============================================================================
🎯  OPTIONAL FLAGS
===============================================================================
--dry-run          → Force skip actual Shopify updates (simulation)
--force            → Update even if product already has a description
--only-sku SKU123  → Process only specific SKU(s); can repeat
--limit 100        → Limit number of discovered products

Examples:
    python linn_to_shopify_desc.py --mode process --dry-run
    python linn_to_shopify_desc.py --mode process --only-sku ABC123 --limit 5
    python linn_to_shopify_desc.py --mode update --force

===============================================================================
📊  TRACKER FILE STRUCTURE (Excel)
===============================================================================
shopify_desc_fill_tracker.xlsx → Sheet "DescFill"

| Timestamp | SKU | ProductID | Title | VariantIDs | CurrentDescLen |
| NewDescLen | Source | SubSource | Status | Note | DryRun |

Statuses include:
    PENDING         - Found but not processed yet
    POPULATED       - Description fetched from Linnworks
    POPULATE_READY  - Populated but dry-run (not written)
    UPDATED         - Successfully written to Shopify
    VERIFIED        - Shopify confirmed new description
    MISSING_DESC    - Linnworks description empty
    NOT_FOUND       - SKU not found in Linnworks
    ERROR           - API or other error
    VERIFY_FAIL     - Shopify still empty after update
    SKIPPED         - Product already had description

===============================================================================
💡  NOTES
===============================================================================
• Shopify body_html is replaced only if blank, unless --force used.
• Linnworks HTML is uploaded as-is (no text cleaning/modification).
• All steps can be safely stopped/restarted — results are appended in tracker.
• DRY_RUN=true ensures nothing is written to Shopify (good for testing).
• Recommended safe speeds:
      REQUEST_DELAY=0.2  (Linnworks)
      SHOPIFY_REQUEST_DELAY=0.5  (Shopify)
• Always check the tracker Excel after each run to review results.

===============================================================================
🧑‍💻  AUTHOR:  Raju / Hiren (Shatchi/Frono Automation)
===============================================================================
"""

import os, sys, time, json, argparse, math, traceback
from typing import Any, Dict, List, Tuple, Optional
from datetime import datetime
import requests
import re
from dotenv import load_dotenv, find_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# =======================
# ENV & CONFIG
# =======================
def load_env():
    # Respect explicit path if provided, else auto-discover
    env_path = os.getenv("ENV_PATH")
    if env_path and os.path.exists(env_path):
        load_dotenv(env_path)
    else:
        load_dotenv(find_dotenv(usecwd=True))

load_env()

# Linnworks
APP_ID      = (os.getenv("LINNWORKS_APPLICATION_ID") or "").strip()
APP_SECRET  = (os.getenv("LINNWORKS_APPLICATION_SECRET") or "").strip()
GRANT_TOKEN = (os.getenv("LINNWORKS_GRANT_TOKEN") or "").strip()

# Shopify (REST)
SHOPIFY_STORE_NAME  = (os.getenv("SHOPIFY_STORE_NAME") or "").strip()      # e.g., myshop (not full domain)
SHOPIFY_ACCESS_TOKEN= (os.getenv("SHOPIFY_ACCESS_TOKEN") or "").strip()
API_VERSION         = (os.getenv("API_VERSION") or "2025-01").strip()

# Behaviour
DRY_RUN             = (os.getenv("DRY_RUN","false").lower() in ("1","true","yes"))
FORCE_UPDATE        = (os.getenv("FORCE_UPDATE","false").lower() in ("1","true","yes"))

# Channels (optional, defaults)
CHANNEL_SOURCE      = (os.getenv("CHANNEL_SOURCE") or "EBAY").strip()
CHANNEL_SUBSOURCE   = (os.getenv("CHANNEL_SUBSOURCE") or "EBAY1_UK").strip()

# Pacing
REQUEST_DELAY            = float(os.getenv("REQUEST_DELAY", "0.2"))           # Linnworks
SHOPIFY_REQUEST_DELAY    = float(os.getenv("SHOPIFY_REQUEST_DELAY", "0.5"))   # Shopify

# Files
TRACKER_PATH = os.getenv("TRACKER_PATH", "shopify_desc_fill_tracker.xlsx")
STATE_PATH   = os.getenv("STATE_PATH", ".linnworks_shopify_desc.state.json")

# Additional options
ADD_DESC_HEADER      = (os.getenv("ADD_DESC_HEADER","true").lower() in ("1","true","yes"))
WRAP_PLAIN_TEXT_HTML = (os.getenv("WRAP_PLAIN_TEXT","true").lower() in ("1","true","yes"))
# =======================

HEADERS = [
    "Timestamp","SKU","ProductID","Title","VariantIDs","CurrentDescLen",
    "NewDescLen","Source","SubSource","Status","Note","DryRun"
]
STATUSES = {
    "PENDING","POPULATED","POPULATE_READY","UPDATED","VERIFIED",
    "MISSING_DESC","NOT_FOUND","ERROR","VERIFY_FAIL","SKIPPED"
}

# =======================
# UTIL
# =======================
def ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def ensure_tracker():
    if not os.path.exists(TRACKER_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "DescFill"
        ws.append(HEADERS)
        # basic column sizing
        widths = [20,25,15,40,20,16,12,10,12,16,60,8]
        for i,w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(TRACKER_PATH)

def append_rows(rows: List[List[Any]]):
    ensure_tracker()
    wb = load_workbook(TRACKER_PATH)
    ws = wb["DescFill"]
    for r in rows:
        ws.append(r)
    wb.save(TRACKER_PATH)

def load_state() -> Dict[str, Any]:
    if os.path.exists(STATE_PATH):
        with open(STATE_PATH,"r",encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_state(state: Dict[str, Any]):
    with open(STATE_PATH,"w",encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def sleep_s(t: float):
    if t > 0:
        time.sleep(t)

def log(msg: str):
    print(msg, flush=True)

def ensure_description_header_and_html(html: str) -> str:
    """
    - If content is plain text, wrap into <p> and convert newlines to <br>.
    - If 'Description:' header is not present near the start, prepend it.
    """
    if not html:
        return html
    clean = html.strip()
    lower = clean.lower()

    # Detect if it's already HTML-ish
    looks_html = any(tag in lower for tag in ("<p", "<br", "<ul", "<ol", "<div", "<span", "<table"))

    # Wrap plain text
    if WRAP_PLAIN_TEXT_HTML and not looks_html:
        # normalize Windows newlines too
        clean = clean.replace("\r\n", "\n")
        replaced = clean.replace("\n", "<br>")
        clean = "<p>" + replaced + "</p>"
        lower = clean.lower()

    # Ensure Description: header near the start (case-insensitive)
    if ADD_DESC_HEADER:
        # Check the first ~80 chars (after stripping) for 'description:'
        # This is lenient: if it already exists, don't add again.
        snippet = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", clean)).strip().lower()[:80]
        if "description:" not in snippet:
            clean = f"<p><strong>Description:</strong></p>\n{clean}"

    return clean


# =======================
# LINNWORKS
# =======================
def lw_authorize() -> Tuple[str, str]:
    if not (APP_ID and APP_SECRET and GRANT_TOKEN):
        raise SystemExit("Missing Linnworks .env vars LINNWORKS_APPLICATION_*")
    url = "https://api.linnworks.net/api/Auth/AuthorizeByApplication"
    payload = {"ApplicationId": APP_ID, "ApplicationSecret": APP_SECRET, "Token": GRANT_TOKEN}
    r = requests.post(url, json=payload, timeout=40)
    r.raise_for_status()
    data = r.json()
    token  = data.get("Token")
    server = (os.getenv("LINNWORKS_SERVER_OVERRIDE") or
              data.get("Server") or data.get("ServerAddress") or data.get("ServerUrl") or "").rstrip("/")
    if not token or not server:
        raise RuntimeError(f"Linnworks auth response missing Token/Server: {data}")
    return token, server

def lw_make_session_with_probe():
    # Auth
    r = requests.post("https://api.linnworks.net/api/Auth/AuthorizeByApplication",
                      json={"ApplicationId":APP_ID,"ApplicationSecret":APP_SECRET,"Token":GRANT_TOKEN}, timeout=40)
    r.raise_for_status()
    d = r.json()
    token  = d["Token"]
    server = (os.getenv("LINNWORKS_SERVER_OVERRIDE") or d.get("Server") or "").rstrip("/")
    if not server:
        raise SystemExit("No Linnworks server from auth; set LINNWORKS_SERVER_OVERRIDE")

    def probe(style):
        h = {"Authorization": (f"Bearer {token}" if style=="Bearer" else token),
             "Accept":"application/json","Content-Type":"application/json"}
        pr = requests.get(f"{server}/api/Inventory/GetChannels", headers=h, timeout=20)
        return pr.status_code==200, h

    forced = (os.getenv("LINNWORKS_AUTH_STYLE") or "").strip()
    if forced in ("Bearer","Raw"):
        ok, h = probe(forced)
        if not ok: raise SystemExit(f"Probe failed with forced style {forced} on {server}")
        s = requests.Session(); s.headers.update(h)
        return s, server, forced

    ok, h = probe("Bearer")
    if ok:
        s = requests.Session(); s.headers.update(h)
        return s, server, "Bearer"
    ok, h = probe("Raw")
    if ok:
        s = requests.Session(); s.headers.update(h)
        return s, server, "Raw"
    raise SystemExit("Auth failed with both Bearer and Raw on " + server)



def lw_session(auth: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {auth}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    })
    return s

def lw_post(s: requests.Session, server: str, path: str, payload: Dict[str,Any]) -> Any:
    url = f"{server}/api{path}" if not server.endswith("/api") else f"{server}{path}"
    r = s.post(url, data=json.dumps(payload), timeout=50)
    if r.status_code != 200:
        raise RuntimeError(f"LW HTTP {r.status_code} on {path}: {r.text[:300]}")
    return r.json()

def lw_get(s: requests.Session, server: str, path: str, params: Dict[str,Any]) -> Any:
    url = f"{server}/api{path}" if not server.endswith("/api") else f"{server}{path}"
    r = s.get(url, params=params, timeout=50)
    if r.status_code != 200:
        raise RuntimeError(f"LW HTTP {r.status_code} on {path}: {r.text[:300]}")
    return r.json()

def lw_get_stock_ids(s: requests.Session, server: str, skus: List[str]) -> Dict[str,str]:
    mapping={}
    try:
        data = lw_post(s, server, "/Inventory/GetStockItemIdsBySKU", {"request":{"SKUS": skus}})
        for it in (data or {}).get("Items", []):
            sku, sid = it.get("SKU"), it.get("StockItemId")
            if sku and sid: mapping[sku]=sid
        return mapping
    except Exception:
        # fallback content-type
        url = f"{server}/api/Inventory/GetStockItemIdsBySKU"
        headers = dict(s.headers); headers["Content-Type"] = "application/x-www-form-urlencoded"
        r = s.post(url, data={"request": json.dumps({"SKUS": skus})}, headers=headers, timeout=50)
        if r.status_code != 200:
            raise RuntimeError(f"LW fallback HTTP {r.status_code}: {r.text[:300]}")
        data = r.json()
        for it in (data or {}).get("Items", []):
            sku, sid = it.get("SKU"), it.get("StockItemId")
            if sku and sid: mapping[sku]=sid
        return mapping

def lw_get_desc_rows(s: requests.Session, server: str, stock_item_id: str) -> List[Dict[str,Any]]:
    return lw_get(s, server, "/Inventory/GetInventoryItemDescriptions", {"inventoryItemId": stock_item_id})

def lw_pick_channel_desc(desc_rows: List[Dict[str,Any]], source: str, subsource: str) -> str:
    for d in desc_rows or []:
        if (d.get("Source") or "").upper() == source.upper() and (d.get("SubSource") or "") == subsource:
            return d.get("Description") or ""
    return ""

# =======================
# SHOPIFY
# =======================
def shopify_base() -> str:
    if not SHOPIFY_STORE_NAME or not SHOPIFY_ACCESS_TOKEN:
        raise SystemExit("Missing Shopify .env vars SHOPIFY_STORE_NAME / SHOPIFY_ACCESS_TOKEN")
    return f"https://{SHOPIFY_STORE_NAME}.myshopify.com/admin/api/{API_VERSION}"

def shopify_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Accept": "application/json",
        "Content-Type": "application/json",
    })
    return s

def sh_get(s: requests.Session, path: str, params: Dict[str,Any]=None) -> requests.Response:
    url = f"{shopify_base()}{path}"
    return s.get(url, params=params or {}, timeout=50)

def sh_put(s: requests.Session, path: str, payload: Dict[str,Any]) -> requests.Response:
    url = f"{shopify_base()}{path}"
    return s.put(url, json=payload, timeout=50)

def sh_paginate_products_empty_desc(s: requests.Session, limit: int=250):
    """Yield products with empty/whitespace body_html, pages via link headers."""
    params = {"limit": limit, "fields":"id,title,body_html,variants"}
    path = "/products.json"
    next_url = None
    while True:
        if next_url:
            r = s.get(next_url, timeout=50)
        else:
            r = sh_get(s, path, params)
        if r.status_code == 429:
            time.sleep(2);  # backoff
            continue
        r.raise_for_status()
        data = r.json()
        prods = data.get("products", [])
        for p in prods:
            body = (p.get("body_html") or "").strip()
            if body == "":
                yield p
        # pagination
        link = r.headers.get("Link") or r.headers.get("link")
        if not link or 'rel="next"' not in link:
            break
        # parse page_info next
        # Link: <https://.../products.json?limit=250&page_info=...>; rel="next"
        parts = [x.strip() for x in link.split(",")]
        next_url = None
        for part in parts:
            if 'rel="next"' in part:
                left = part.split(";")[0].strip()
                if left.startswith("<") and left.endswith(">"):
                    next_url = left[1:-1]
        if not next_url:
            break

def sh_get_product(s: requests.Session, product_id: int) -> Dict[str,Any]:
    r = sh_get(s, f"/products/{product_id}.json")
    if r.status_code == 429:
        time.sleep(2); return sh_get_product(s, product_id)
    r.raise_for_status()
    return r.json().get("product", {})

def sh_update_product_desc(s: requests.Session, product_id: int, html: str) -> None:
    payload = {"product": {"id": product_id, "body_html": html}}
    r = sh_put(s, f"/products/{product_id}.json", payload)
    if r.status_code == 429:
        time.sleep(2); return sh_update_product_desc(s, product_id, html)
    if r.status_code >= 400:
        raise RuntimeError(f"Shopify PUT {r.status_code}: {r.text[:300]}")
    return

# =======================
# DISCOVER → candidates
# =======================
def discover_candidates(limit: Optional[int]=None, only_skus: Optional[List[str]]=None) -> List[Dict[str,Any]]:
    log("[DISCOVER] Scanning Shopify products with empty body_html…")
    s = shopify_session()
    out=[]
    count=0
    for p in sh_paginate_products_empty_desc(s):
        # collect SKUs from variants
        skus = [ (v.get("sku") or "").strip() for v in (p.get("variants") or []) ]
        skus = [x for x in skus if x]
        if not skus:
            # no SKU to resolve Linnworks against — still track as SKIPPED/PENDING
            row = {
                "Timestamp": ts(),
                "SKU": "",
                "ProductID": p["id"],
                "Title": p.get("title") or "",
                "VariantIDs": ",".join([str(v.get("id")) for v in p.get("variants",[])]),
                "CurrentDescLen": 0,
                "NewDescLen": 0,
                "Source": CHANNEL_SOURCE,
                "SubSource": CHANNEL_SUBSOURCE,
                "Status": "PENDING",
                "Note": "No SKU on any variants",
                "DryRun": DRY_RUN
            }
            out.append(row)
            count += 1
        else:
            for sku in skus:
                if only_skus and sku not in only_skus:
                    continue
                row = {
                    "Timestamp": ts(),
                    "SKU": sku,
                    "ProductID": p["id"],
                    "Title": p.get("title") or "",
                    "VariantIDs": ",".join([str(v.get("id")) for v in p.get("variants",[])]),
                    "CurrentDescLen": 0,
                    "NewDescLen": 0,
                    "Source": CHANNEL_SOURCE,
                    "SubSource": CHANNEL_SUBSOURCE,
                    "Status": "PENDING",
                    "Note": "",
                    "DryRun": DRY_RUN
                }
                out.append(row)
                count += 1
        if limit and count >= limit:
            break
        sleep_s(SHOPIFY_REQUEST_DELAY)
    log(f"[DISCOVER] Found {len(out)} candidate rows.")
    return out

# =======================
# POPULATE → Linnworks HTML
# =======================
def chunked(seq: List[str], n: int):
    for i in range(0, len(seq), n):
        yield seq[i:i+n]

def populate_from_linnworks(pending_rows: List[Dict[str,Any]]) -> List[Dict[str,Any]]:
    log("[POPULATE] Fetching descriptions from Linnworks…")
    # make SKUs set
    want_skus = [r["SKU"] for r in pending_rows if r.get("SKU")]
    if not want_skus:
        log("[POPULATE] No SKUs to resolve.")
        return pending_rows

    s, server, auth_style = lw_make_session_with_probe()
    print(f"[LW] Auth style={auth_style} server={server}")


    # resolve SKUs → stock ids in batches
    sku_to_id: Dict[str,str]={}
    for batch in chunked(want_skus, 80):
        mapping = lw_get_stock_ids(s, server, batch)
        sku_to_id.update(mapping)
        sleep_s(REQUEST_DELAY)

    for r in pending_rows:
        sku = r.get("SKU") or ""
        if not sku:
            r["Status"]="SKIPPED"
            r["Note"] = (r.get("Note") or "") + " | No SKU"
            continue
        sid = sku_to_id.get(sku)
        if not sid:
            r["Status"]="NOT_FOUND"
            r["Note"]="SKU not found in Linnworks"
            continue
        try:
            desc_rows = lw_get_desc_rows(s, server, sid)
            html = lw_pick_channel_desc(desc_rows, r["Source"], r["SubSource"])
            if html.strip():
                prepared = ensure_description_header_and_html(html)
                r["NewDescLen"] = len(prepared)
                r["_NewHTML"] = prepared  # temp
                r["Status"] = "POPULATED" if not DRY_RUN else "POPULATE_READY"
            else:
                r["Status"] = "MISSING_DESC"
                r["Note"]   = "Channel description empty"

        except Exception as e:
            r["Status"]="ERROR"
            r["Note"]=f"LW error: {e}"
        sleep_s(REQUEST_DELAY)
    return pending_rows

# =======================
# UPDATE → Shopify
# =======================
def update_shopify(rows: List[Dict[str,Any]], force: bool=False) -> List[Dict[str,Any]]:
    log("[UPDATE] Writing product descriptions to Shopify…")
    s = shopify_session()
    for r in rows:
        if r.get("Status") not in {"POPULATED","POPULATE_READY"}:
            continue
        if DRY_RUN:
            # don't write, just mark as would-update
            r["Status"]="POPULATE_READY"
            r["Note"]=(r.get("Note") or "") + " | DRY_RUN: no write"
            continue

        product_id = int(r["ProductID"])
        html = r.get("_NewHTML","") or ""
        if not html.strip():
            r["Status"]="ERROR"; r["Note"]="No HTML to write"; continue

        try:
            # if not forcing, double-check product still empty
            if not force:
                p = sh_get_product(s, product_id)
                cur = (p.get("body_html") or "").strip()
                if cur:
                    r["Status"]="SKIPPED"
                    r["Note"]=(r.get("Note") or "") + " | Product already has description"
                    sleep_s(SHOPIFY_REQUEST_DELAY)
                    continue

            sh_update_product_desc(s, product_id, html)
            r["Status"]="UPDATED"
        except Exception as e:
            r["Status"]="ERROR"
            r["Note"]=f"Shopify update error: {e}"
        sleep_s(SHOPIFY_REQUEST_DELAY)
    return rows

# =======================
# VERIFY → read back
# =======================
def verify_updates(rows: List[Dict[str,Any]]) -> List[Dict[str,Any]]:
    log("[VERIFY] Re-reading Shopify products to confirm…")
    s = shopify_session()
    for r in rows:
        if r.get("Status") not in {"UPDATED","POPULATE_READY"}:
            continue
        if r["Status"] == "POPULATE_READY":  # DRY_RUN case
            # treat as verified preview
            continue
        try:
            p = sh_get_product(s, int(r["ProductID"]))
            cur = (p.get("body_html") or "").strip()
            if cur:
                r["Status"]="VERIFIED"
            else:
                r["Status"]="VERIFY_FAIL"
                r["Note"]=(r.get("Note") or "") + " | Still empty after update"
        except Exception as e:
            r["Status"]="ERROR"
            r["Note"]=f"Verify error: {e}"
        sleep_s(SHOPIFY_REQUEST_DELAY)
    return rows

# =======================
# TRACKER IO
# =======================
def to_rowdicts_for_append(rows: List[Dict[str,Any]]) -> List[List[Any]]:
    out=[]
    for r in rows:
        out.append([
            r.get("Timestamp") or ts(),
            r.get("SKU",""),
            r.get("ProductID",""),
            r.get("Title",""),
            r.get("VariantIDs",""),
            r.get("CurrentDescLen",0),
            r.get("NewDescLen",0),
            r.get("Source",CHANNEL_SOURCE),
            r.get("SubSource",CHANNEL_SUBSOURCE),
            r.get("Status",""),
            r.get("Note",""),
            str(DRY_RUN).lower()
        ])
    return out

# =======================
# MAIN MODES
# =======================
def run_discover(args):
    rows = discover_candidates(limit=args.limit, only_skus=args.only_sku)
    append_rows(to_rowdicts_for_append(rows))
    log(f"[DISCOVER] Wrote {len(rows)} rows to {TRACKER_PATH}")

def run_populate(args):
    # read last sheet and pick PENDING
    ensure_tracker()
    wb = load_workbook(TRACKER_PATH)
    ws = wb["DescFill"]
    header = [c.value for c in ws[1]]
    idx = {name:i+1 for i,name in enumerate(header)}

    pending=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        status = (r[idx["Status"]-1] or "").strip()
        if status == "PENDING":
            pending.append({
                "Timestamp": r[idx["Timestamp"]-1],
                "SKU": r[idx["SKU"]-1] or "",
                "ProductID": r[idx["ProductID"]-1],
                "Title": r[idx["Title"]-1] or "",
                "VariantIDs": r[idx["VariantIDs"]-1] or "",
                "CurrentDescLen": r[idx["CurrentDescLen"]-1] or 0,
                "NewDescLen": r[idx["NewDescLen"]-1] or 0,
                "Source": r[idx["Source"]-1] or CHANNEL_SOURCE,
                "SubSource": r[idx["SubSource"]-1] or CHANNEL_SUBSOURCE,
                "Status": status,
                "Note": r[idx["Note"]-1] or "",
            })
    wb.close()

    if not pending:
        log("[POPULATE] No PENDING rows found.")
        return

    populated = populate_from_linnworks(pending)
    append_rows(to_rowdicts_for_append(populated))
    log(f"[POPULATE] Appended {len(populated)} result rows to {TRACKER_PATH}")

def run_update(args):
    ensure_tracker()
    wb = load_workbook(TRACKER_PATH)
    ws = wb["DescFill"]
    header = [c.value for c in ws[1]]
    idx = {name:i+1 for i,name in enumerate(header)}

    # take latest statuses for each (SKU, ProductID) where POPULATED/POPULATE_READY
    candidates=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        st = (r[idx["Status"]-1] or "").strip()
        if st in {"POPULATED","POPULATE_READY"}:
            candidates.append({
                "Timestamp": r[idx["Timestamp"]-1],
                "SKU": r[idx["SKU"]-1] or "",
                "ProductID": r[idx["ProductID"]-1],
                "Title": r[idx["Title"]-1] or "",
                "VariantIDs": r[idx["VariantIDs"]-1] or "",
                "CurrentDescLen": r[idx["CurrentDescLen"]-1] or 0,
                "NewDescLen": r[idx["NewDescLen"]-1] or 0,
                "Source": r[idx["Source"]-1] or CHANNEL_SOURCE,
                "SubSource": r[idx["SubSource"]-1] or CHANNEL_SUBSOURCE,
                "Status": st,
                "Note": r[idx["Note"]-1] or "",
                # placeholder for html will be missing if you restarted process,
                # but we verified populate step just appended a fresh row with _NewHTML cached in memory only.
            })
    wb.close()

    if not candidates:
        log("[UPDATE] No POPULATED/POPULATE_READY rows found.")
        return

    # We need HTML to update. If the process was restarted, HTML isn’t in tracker.
    # Simple re-populate on-the-fly for the subset:
    candidates = populate_from_linnworks(candidates)

    updated = update_shopify(candidates, force=args.force)
    append_rows(to_rowdicts_for_append(updated))
    log(f"[UPDATE] Appended {len(updated)} result rows to {TRACKER_PATH}")

def run_verify(args):
    ensure_tracker()
    wb = load_workbook(TRACKER_PATH)
    ws = wb["DescFill"]
    header = [c.value for c in ws[1]]
    idx = {name:i+1 for i,name in enumerate(header)}

    to_check=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        st = (r[idx["Status"]-1] or "").strip()
        if st in {"UPDATED","POPULATE_READY"}:
            to_check.append({
                "Timestamp": r[idx["Timestamp"]-1],
                "SKU": r[idx["SKU"]-1] or "",
                "ProductID": r[idx["ProductID"]-1],
                "Title": r[idx["Title"]-1] or "",
                "VariantIDs": r[idx["VariantIDs"]-1] or "",
                "CurrentDescLen": r[idx["CurrentDescLen"]-1] or 0,
                "NewDescLen": r[idx["NewDescLen"]-1] or 0,
                "Source": r[idx["Source"]-1] or CHANNEL_SOURCE,
                "SubSource": r[idx["SubSource"]-1] or CHANNEL_SUBSOURCE,
                "Status": st,
                "Note": r[idx["Note"]-1] or "",
            })
    wb.close()

    if not to_check:
        log("[VERIFY] No UPDATED/POPULATE_READY rows to verify.")
        return

    verified = verify_updates(to_check)
    append_rows(to_rowdicts_for_append(verified))
    log(f"[VERIFY] Appended {len(verified)} result rows to {TRACKER_PATH}")

def run_process(args):
    # 1) discover
    rows = discover_candidates(limit=args.limit, only_skus=args.only_sku)
    append_rows(to_rowdicts_for_append(rows))
    # 2) populate
    populated = populate_from_linnworks(rows)
    append_rows(to_rowdicts_for_append(populated))
    # 3) update
    updated = update_shopify(populated, force=args.force)
    append_rows(to_rowdicts_for_append(updated))
    # 4) verify
    verified = verify_updates(updated)
    append_rows(to_rowdicts_for_append(verified))
    log("[PROCESS] Done. See tracker for statuses.")

# =======================
# CLI
# =======================
def parse_args():
    ap = argparse.ArgumentParser(description="Fill empty Shopify product descriptions from Linnworks channel descriptions.")
    ap.add_argument("--mode", required=True, choices=["discover","populate","update","verify","process"], help="Which step to run")
    ap.add_argument("--only-sku", action="append", help="Limit to specific SKU(s); can be repeated")
    ap.add_argument("--limit", type=int, help="Limit number of discovered products")
    ap.add_argument("--force", action="store_true", help="Overwrite non-empty Shopify descriptions")
    ap.add_argument("--dry-run", action="store_true", help="Override DRY_RUN=true")
    return ap.parse_args()

def main():
    args = parse_args()
    global DRY_RUN
    if args.dry_run:
        DRY_RUN = True
    if args.force:
        # also reflect in env-style flag if user wants
        pass

    log(f"Store: {SHOPIFY_STORE_NAME}  API: {API_VERSION}  DRY_RUN={DRY_RUN}  FORCE={args.force or FORCE_UPDATE}")
    log(f"Channel: {CHANNEL_SOURCE}/{CHANNEL_SUBSOURCE}")
    try:
        if args.mode == "discover":
            run_discover(args)
        elif args.mode == "populate":
            run_populate(args)
        elif args.mode == "update":
            run_update(args)
        elif args.mode == "verify":
            run_verify(args)
        elif args.mode == "process":
            run_process(args)
    except Exception as e:
        log(f"[FATAL] {e}\n{traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    main()
